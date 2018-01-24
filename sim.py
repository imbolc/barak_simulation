import math
from copy import deepcopy
from pathlib import Path

import openpyxl


SRC_FILE = 'drug_delivery_data.xlsx'
DST_FOLDER = 'results'
MOLECULES = ['Acetylcholine', 'Norepinephrine', 'Dopamine', 'Serotonin']


def simulation():
    root = Path(__file__).resolve().with_name(DST_FOLDER)
    root.mkdir(exist_ok=True)

    patients = read_xls()
    for id, patient in patients.items():
        data = []
        data.append((patient, get_softmax(patient)))
        for molecule in patient['molecules'].keys():
            for arg in range(5):
                p = deepcopy(patient)
                p['molecules'][molecule][arg] = 1
                data.append((p, get_softmax(p)))
        save_patient(root / f'patient-{id}.xlsx', data)


def save_patient(path, data):
    book = openpyxl.load_workbook('./tpl.xlsx')
    sheet = book.active
    add_row = add_to_sheet(sheet)
    for case, prob in data:
        for molecule in MOLECULES:
            row = [molecule]
            row.extend(case['molecules'][molecule])
            row.append(prob[molecule])
            add_row(row)
    book.save(path)
    return


def add_to_sheet(sheet):
    row_num = 2

    def add(columns):
        nonlocal row_num
        for i, val in enumerate(columns, start=1):
            sheet.cell(row=row_num, column=i, value=val)
        row_num += 1
    return add


def read_xls():
    patients = {}
    book = openpyxl.load_workbook(SRC_FILE, read_only=True)
    assert len(book.sheetnames) == 1, 'Found multiple sheets in source file'
    sheet = book.active
    for i, row in enumerate(sheet.rows):
        if not i:
            continue
        id, in_use, molecule, *intensities = [r.value for r in row]
        intensities = intensities[:5]
        assert len(intensities) == 5
        patient = patients.setdefault(int(id), {})
        if molecule == 'weights':
            patient['weights'] = intensities
        else:
            assert molecule in MOLECULES, molecule
            patient.setdefault('molecules', {})[molecule] = intensities
    return patients


def get_softmax(patient):
    weights = patient['weights']
    weighted = {}
    for molecule, intensities in patient['molecules'].items():
        weighted[molecule] = sum(a * w for a, w in zip(intensities, weights))
    denom = sum(math.exp(k) for k in weighted.values())
    return {k: math.exp(v) / denom for k, v in weighted.items()}


if __name__ == '__main__':
    simulation()
